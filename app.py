import os
import socket
from urllib.parse import urlparse, urlunparse
from flask import request
import time
import secrets
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import calendar
from functools import wraps
from typing import Tuple

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

BASE_URL = os.environ.get("BASE_URL")  # Optional public base URL; commonly unset for local-only runs

QR_DEFAULT_INTERVAL = 90  # default QR refresh interval in seconds

SUBJECTS = {
    "MATH01": {"name": "Mathematics", "password": "math2025"},
    "PHY01": {"name": "Physics", "password": "phys2025"},
    "CHEM01": {"name": "Chemistry", "password": "chem2025"},
    "BIO01": {"name": "Biology", "password": "bio2025"},
    "ENG01": {"name": "English", "password": "eng2025"},
    "COMP01": {"name": "Computer Science", "password": "comp2025"},
    "HIST01": {"name": "History", "password": "hist2025"},
    "GEO01": {"name": "Geography", "password": "geo2025"},
    "ECO01": {"name": "Economics", "password": "eco2025"},
    "STAT01": {"name": "Statistics", "password": "stat2025"},
}

EXCEL_FILE = "Attendance.xlsx"
ATTENDANCE_THRESHOLD = 75

active_sessions = {}  # {subject_code: {'token': ..., 'generated': ..., 'interval': ..., 'history': [...]}}

# Helper functions

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        master = wb.active
        master.title = "Master"
        master.append(
            ["Date", "Subject", "Roll No", "Present", "Token ID", "Timestamp", "IP Address"]
        )
        summary = wb.create_sheet("Summary")
        summary.append(
            [
                "Week Ending",
                "Roll No",
                "Subject",
                "Classes Conducted",
                "Classes Attended",
                "Attendance Percentage",
                "Status",
            ]
        )
        suspicious = wb.create_sheet("Suspicious")
        suspicious.append([
            "Date",
            "Subject",
            "Roll No",
            "IP Address",
            "Token ID",
            "Timestamp",
            "Notes",
        ])
        wb.save(EXCEL_FILE)


def get_or_create_sheet(wb, subject_code: str, date_str: str) -> Tuple[bool, object]:
    sheet_name = f"{subject_code}_{date_str}"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Roll No", "Present", "Token ID", "Timestamp", "IP Address"])
        for r in range(1, 81):
            ws.append([r, "Absent", "", "", ""])
        wb.save(EXCEL_FILE)
        return True, ws
    return False, wb[sheet_name]


def update_summary_sheet(wb):
    today = date.today()
    if today.weekday() != calendar.SATURDAY:
        return
    summary = wb["Summary"]
    sheets = [s for s in wb.sheetnames if any(s.startswith(code) for code in SUBJECTS.keys())]
    attendance_data = {}
    for sheet_name in sheets:
        scode = sheet_name.split("_")[0]
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll = row[0]
            present = row[1] == "Present"
            if roll not in attendance_data:
                attendance_data[roll] = {c: {"attended": 0, "total": 0} for c in SUBJECTS}
            attendance_data[roll][scode]["total"] += 1
            if present:
                attendance_data[roll][scode]["attended"] += 1
    week_end = today.strftime("%Y-%m-%d")
    for roll, subjects in attendance_data.items():
        for scode, counts in subjects.items():
            if counts["total"] > 0:
                perc = counts["attended"] / counts["total"] * 100
                status = "Satisfactory" if perc >= ATTENDANCE_THRESHOLD else "Below Required"
                summary.append(
                    [
                        week_end,
                        roll,
                        scode,
                        counts["total"],
                        counts["attended"],
                        f"{perc:.2f}%",
                        status,
                    ]
                )
    wb.save(EXCEL_FILE)

def get_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip


def ensure_reachable_url(url: str) -> str:
    """If URL points to localhost/127.0.0.1, replace hostname with LAN IP.

    This helps when generating QR codes on a machine so mobile devices on the same
    Wi-Fi can reach the server using its LAN address.
    """
    try:
        parsed = urlparse(url)
        hostname = parsed.hostname
        if hostname in ("127.0.0.1", "localhost"):
            lan_ip = get_ip()
            # preserve port if present
            port = parsed.port
            netloc = f"{lan_ip}:{port}" if port else lan_ip
            new = parsed._replace(netloc=netloc)
            return urlunparse(new)
    except Exception:
        pass
    return url


def get_qr_interval() -> int:
    try:
        interval = int(session.get("qr_interval", QR_DEFAULT_INTERVAL))
        if interval in [30, 45, 60, 90]:
            return interval
    except Exception:
        pass
    return QR_DEFAULT_INTERVAL


def generate_token(subject_code: str) -> str:
    token = secrets.token_urlsafe(16)[:16]
    interval = get_qr_interval()
    history = active_sessions.get(subject_code, {}).get("history", [])
    history.append({"token": token, "generated": time.time(), "interval": interval})
    active_sessions[subject_code] = {
        "token": token,
        "generated": time.time(),
        "interval": interval,
        "history": history,
    }
    return token


def generate_subject_qr(subject_code: str) -> Tuple[str, str]:
    if subject_code not in SUBJECTS:
        raise ValueError("Invalid subject code")
    token = generate_token(subject_code)
    now = int(time.time())
    base = BASE_URL
    if base:
        url = f"{base.rstrip('/')}/attend/{token}?t={now}"
    else:
        # Try to build an external URL from Flask; if not available (e.g., outside request
        # context), fall back to localhost so local testing works.
        try:
            url = url_for('attend', token=token, _external=True) + f"?t={now}"
        except Exception:
                url = f"http://localhost:5000/attend/{token}?t={now}"
        # Ensure QR uses a reachable host (replace localhost with LAN IP if needed)
        url = ensure_reachable_url(url)
    import qrcode

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    qr_dir = os.path.join("static", "qr_codes")
    os.makedirs(qr_dir, exist_ok=True)
    filename = f"qr_{subject_code}_{now}.png"
    path = os.path.join(qr_dir, filename)
    img.save(path)
    latest_path = os.path.join(qr_dir, "latest.png")
    img.save(latest_path)
    return url, "qr_codes/latest.png"


def validate_attendance_session(token: str, subject_code: str) -> Tuple[bool, str]:
    data = active_sessions.get(subject_code)
    if not data or data.get("token") != token:
        return False, "Invalid or expired session token."
    interval = data.get("interval", QR_DEFAULT_INTERVAL)
    if time.time() - data.get("generated", 0) > interval:
        return False, "Attendance session expired."
    return True, ""


def mark_attendance(wb, roll_no: int, token: str, ip_address: str, subject_code: str) -> Tuple[bool, str]:
    valid, msg = validate_attendance_session(token, subject_code)
    if not valid:
        return False, msg
    today_s = date.today().strftime("%Y-%m-%d")
    _, ws = get_or_create_sheet(wb, subject_code, today_s)
    try:
        roll_int = int(roll_no)
        if roll_int < 1 or roll_int > 80:
            return False, "Roll number must be between 1 and 80."
    except Exception:
        return False, "Invalid roll number."
    for row in ws.iter_rows(min_row=2):
        if row[0].value == roll_int:
            if row[1].value == "Present":
                return False, "Attendance already marked."
            row[1].value = "Present"
            row[2].value = token
            row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row[4].value = ip_address
            master = wb["Master"]
            master.append(
                [
                    today_s,
                    subject_code,
                    roll_int,
                    "Present",
                    token,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ip_address,
                ]
            )
            # --- Suspicious IP detection ---
            # If another student in the same session sheet already marked Present with the
            # same IP address, add both (existing + current) to the Suspicious sheet for
            # faculty validation. Avoid duplicate entries in Suspicious.
            def get_or_create_suspicious(wbb):
                if "Suspicious" in wbb.sheetnames:
                    return wbb["Suspicious"]
                ss = wbb.create_sheet("Suspicious")
                ss.append([
                    "Date",
                    "Subject",
                    "Roll No",
                    "IP Address",
                    "Token ID",
                    "Timestamp",
                    "Notes",
                ])
                return ss

            def suspicious_has_entry_by_roll(ssheet, date_s, subj, roll):
                """Check by date+subject+roll to avoid missing entries due to token differences."""
                for r in ssheet.iter_rows(min_row=2, values_only=True):
                    try:
                        if str(r[0]) == str(date_s) and str(r[1]) == str(subj) and str(r[2]) == str(roll):
                            return True
                    except Exception:
                        continue
                return False

            suspicious_sheet = get_or_create_suspicious(wb)
            matched_other = False
            # normalize current IP for comparison
            cur_ip = (ip_address or "").strip()
            for other in ws.iter_rows(min_row=2):
                try:
                    other_roll = other[0].value
                    other_present = other[1].value == "Present"
                    other_ip = (other[4].value or "").strip()
                    other_token = other[2].value
                    other_ts = other[3].value
                except Exception:
                    continue
                if other_present and other_roll != roll_int and other_ip and other_ip == cur_ip:
                    matched_other = True
                    # append existing student's record if not already present (by roll)
                    if not suspicious_has_entry_by_roll(suspicious_sheet, today_s, subject_code, other_roll):
                        suspicious_sheet.append([
                            today_s,
                            subject_code,
                            other_roll,
                            other_ip,
                            other_token or "",
                            other_ts or "",
                            "Duplicate IP detected",
                        ])
            # If any other student matched, ensure current student's record is also logged once
            if matched_other and not suspicious_has_entry_by_roll(suspicious_sheet, today_s, subject_code, roll_int):
                suspicious_sheet.append([
                    today_s,
                    subject_code,
                    roll_int,
                    cur_ip,
                    token or "",
                    row[3].value or "",
                    "Duplicate IP detected",
                ])
            if date.today().weekday() == calendar.SATURDAY:
                update_summary_sheet(wb)
            wb.save(EXCEL_FILE)
            return True, f"Attendance marked successfully for {SUBJECTS[subject_code]['name']}!"
    return False, "Roll number not found."


def login_required(func):
    @wraps(func)
    def decorated(*args, **kwargs):
        if "subject_code" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return decorated


# --- Routes ---


@app.route("/")
def index():
    if "subject_code" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html", subjects=SUBJECTS)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        subject_code = request.form.get("subject_code")
        password = request.form.get("password")
        if subject_code in SUBJECTS and password == SUBJECTS[subject_code]["password"]:
            session["subject_code"] = subject_code
            session["qr_interval"] = QR_DEFAULT_INTERVAL
            return redirect(url_for("dashboard"))
        return render_template("login.html", subjects=SUBJECTS, error="Invalid credentials")
    return render_template("login.html", subjects=SUBJECTS)


@app.route("/logout")
def logout():
    session.pop("subject_code", None)
    session.pop("qr_interval", None)
    return redirect(url_for("login"))


@app.route("/set_interval", methods=["POST"])
@login_required
def set_interval():
    try:
        interval = int(request.form.get("interval", QR_DEFAULT_INTERVAL))
        if interval in [30, 45, 60, 90]:
            session["qr_interval"] = interval
    except Exception:
        session["qr_interval"] = QR_DEFAULT_INTERVAL
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@login_required
def dashboard():
    init_excel()
    subject_code = session["subject_code"]
    interval = get_qr_interval()
    session_data = active_sessions.get(subject_code)
    now = time.time()
    if session_data and (now - session_data["generated"] < interval):
        remaining = int(interval - (now - session_data["generated"]))
        base = BASE_URL
        if base:
            qr_url = f"{base.rstrip('/')}/attend/{session_data['token']}?t={int(session_data['generated'])}"
        else:
            try:
                qr_url = url_for('attend', token=session_data['token'], _external=True) + f"?t={int(session_data['generated'])}"
            except Exception:
                    qr_url = f"http://localhost:5000/attend/{session_data['token']}?t={int(session_data['generated'])}"
            qr_url = ensure_reachable_url(qr_url)
        qr_filename = "qr_codes/latest.png"
    else:
        try:
            qr_url, qr_filename = generate_subject_qr(subject_code)
            remaining = interval
        except Exception:
            qr_url = None
            qr_filename = None
            remaining = 0
    now_ts = int(time.time())
    return render_template(
        "dashboard.html",
        subject_code=subject_code,
        subject_name=SUBJECTS[subject_code]["name"],
        time_remaining=remaining,
        qr_filename=qr_filename,
        qr_url=qr_url,
        interval=interval,
        now_ts=now_ts,
    )


@app.route("/start_session")
@login_required
def start_session():
    subject_code = session["subject_code"]
    interval = get_qr_interval()
    try:
        qr_url, qr_filename = generate_subject_qr(subject_code)
        base = BASE_URL
        if base:
            qr_static = f"{base.rstrip('/')}/static/{qr_filename}"
            attend_url = qr_url
        else:
            qr_static = url_for('static', filename=qr_filename, _external=True)
            attend_url = qr_url
            # make sure links are reachable from other devices on the LAN
            qr_static = ensure_reachable_url(qr_static)
            attend_url = ensure_reachable_url(attend_url)
        return jsonify(
            {
                "status": "success",
                "qr_url": qr_url,
                "attend_url": attend_url,
                "time_remaining": interval,
                "qr_filename": qr_filename,
                "qr_static": qr_static,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# NOTE: removed /set_base_url helper to return repository to local-only behavior


@app.route("/attend/<token>", methods=["GET", "POST"])
def attend(token):
    subject_code = None
    for scode, sdata in active_sessions.items():
        if sdata.get("token") == token:
            subject_code = scode
            break
    if not subject_code:
        return render_template("error.html", error="Invalid or expired token.")
    valid, msg = validate_attendance_session(token, subject_code)
    if not valid:
        return render_template("error.html", error=msg)
    if request.method == "POST":
        roll_no = request.form.get("roll_no")
        ip_addr = request.headers.get("X-Forwarded-For", request.remote_addr)
        if not roll_no:
            return render_template(
                "attend.html",
                token=token,
                error="Please provide your roll number",
                subject_name=SUBJECTS[subject_code]["name"],
            )
        wb = load_workbook(EXCEL_FILE)
        success, msg = mark_attendance(wb, roll_no, token, ip_addr, subject_code)
        if success:
            return render_template("success.html", message=msg)
        else:
            return render_template(
                "attend.html", token=token, error=msg, subject_name=SUBJECTS[subject_code]["name"]
            )
    return render_template(
        "attend.html", token=token, error=None, subject_name=SUBJECTS[subject_code]["name"]
    )
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
